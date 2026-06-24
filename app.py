import streamlit as st
import zipfile
import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="HR Timesheet Generator",
    page_icon="🗂️",
    layout="centered"
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 12px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .main-header h1 { margin: 0; font-size: 2rem; }
    .main-header p  { margin: 0.5rem 0 0; opacity: 0.85; font-size: 1rem; }

    .step-box {
        background: #f8f9ff;
        border: 1px solid #e0e4ff;
        border-left: 4px solid #667eea;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin-bottom: 1rem;
    }
    .step-box h3 { margin: 0 0 0.3rem; color: #4a5568; font-size: 1rem; }
    .step-box p  { margin: 0; color: #718096; font-size: 0.85rem; }

    .emp-card {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        margin-bottom: 1rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }

    .success-banner {
        background: #f0fff4;
        border: 1px solid #9ae6b4;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
        color: #276749;
        font-weight: 500;
    }

    .stButton>button {
        border-radius: 8px;
        font-weight: 500;
    }
    div[data-testid="stExpander"] {
        border: 1px solid #e2e8f0;
        border-radius: 10px;
    }
</style>
""", unsafe_allow_html=True)

# ── Helper functions ──────────────────────────────────────────────────────────

def get_working_days(year: int, month: int) -> list[date]:
    """Return all Mon–Fri dates for the given month."""
    days = []
    _, total = calendar.monthrange(year, month)
    for d in range(1, total + 1):
        dt = date(year, month, d)
        if dt.weekday() < 5:          # 0=Mon … 4=Fri
            days.append(dt)
    return days


def parse_leave_dates(text: str) -> list[date]:
    """Parse comma-separated DD/MM/YYYY dates from a text string."""
    parsed = []
    if not text.strip():
        return parsed
    for token in text.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            parts = token.split("/")
            if len(parts) != 3:
                raise ValueError
            parsed.append(date(int(parts[2]), int(parts[1]), int(parts[0])))
        except Exception:
            st.warning(f"⚠️  Could not read date: **{token}**  — please use DD/MM/YYYY format.")
    return parsed


def extract_employee_names(zip_bytes: bytes) -> list[str]:
    """
    Read folder names inside the zip to find employee names.
    Zip structure:  Timesheet/Leave/<Name>/  and  Timesheet/Non Leave/<Name>/
    Falls back to any folder 2-levels deep if that pattern isn't found.
    """
    names = set()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for info in z.infolist():
            parts = [p for p in info.filename.split("/") if p]
            # Expected depth: Timesheet / Leave or Non Leave / Employee Name / file.png
            if len(parts) >= 3:
                candidate = parts[-2] if info.is_dir() else parts[-2]
                skip = {"Timesheet", "Leave", "Non Leave", "__MACOSX"}
                if candidate not in skip and not candidate.startswith("."):
                    names.add(candidate)
    return sorted(names)


def generate_excel(employee_name: str, working_days: list[date],
                   leave_dates: set[date]) -> bytes:
    """
    Build a .xlsx file matching the Sonali Dutta template exactly:
    Columns:  Resource Name  |  Timesheet Date  |  Client Hours  |  status
    - 8 hours on working days, 0 on leave days
    - Dates formatted as DD/MM/YYYY
    - Status = Approved for every row
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Resource Name ", "Timesheet Date", "Client Hours", "status"]
    header_fill   = PatternFill("solid", fgColor="4472C4")
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="EEF2FF")   # light purple stripe
    date_fmt = "DD/MM/YYYY"
    name_font = Font(name="Arial", size=10)

    for row_idx, work_date in enumerate(working_days, 2):
        hours  = 0 if work_date in leave_dates else 8
        is_alt = (row_idx % 2 == 0)

        values = [employee_name, work_date, hours, "Approved"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = name_font
            cell.alignment = center
            cell.border    = border
            if is_alt:
                cell.fill = alt_fill
            # Format date column
            if col == 2:
                cell.number_format = date_fmt
            # Colour zero-hour rows amber so leaves stand out
            if col == 3 and val == 0:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
                cell.font = Font(name="Arial", size=10, bold=True, color="856404")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Session-state defaults ────────────────────────────────────────────────────
for key in ("step", "zip_bytes", "employees", "leaves", "month", "year"):
    if key not in st.session_state:
        st.session_state[key] = None

if st.session_state.step is None:
    st.session_state.step = 1
if st.session_state.leaves is None:
    st.session_state.leaves = {}


# ── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>🗂️ HR Timesheet Generator</h1>
  <p>Upload employee zip → Add leave dates → Download Excel files ready to upload</p>
</div>
""", unsafe_allow_html=True)

# ── STEP INDICATOR ───────────────────────────────────────────────────────────
col1, col2, col3, col4 = st.columns(4)
steps = [
    ("1", "Upload Zip",    col1),
    ("2", "Set Leaves",    col2),
    ("3", "Review",        col3),
    ("4", "Download",      col4),
]
for num, label, col in steps:
    active = st.session_state.step == int(num)
    col.markdown(
        f"<div style='text-align:center;padding:8px;border-radius:8px;"
        f"background:{'#667eea' if active else '#f0f0f0'};"
        f"color:{'white' if active else '#888'};font-size:0.8rem;font-weight:600'>"
        f"<div style='font-size:1.1rem'>{'●' if active else '○'}</div>"
        f"Step {num}<br><small>{label}</small></div>",
        unsafe_allow_html=True
    )

st.write("")  # spacer

# ═════════════════════════════════════════════════════════════════════════════
# STEP 1 — Upload zip
# ═════════════════════════════════════════════════════════════════════════════
if st.session_state.step == 1:
    st.markdown("""
    <div class="step-box">
      <h3>📁 Step 1 — Upload the employee timesheet zip file</h3>
      <p>The zip should have folders with each employee's name containing their weekly screenshot images.</p>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Choose your zip file",
        type=["zip"],
        help="Example zip structure:  Timesheet/Leave/Kedar/ and Timesheet/Non Leave/Sanket/"
    )

    col_m, col_y = st.columns(2)
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    with col_m:
        month_label = st.selectbox("Timesheet Month", month_names, index=4)
        chosen_month = month_names.index(month_label) + 1
    with col_y:
        chosen_year = st.number_input("Year", min_value=2020, max_value=2035, value=2026, step=1)

    if st.button("➡️  Detect Employees & Continue", use_container_width=True, type="primary"):
        if uploaded is None:
            st.error("Please upload a zip file first.")
        else:
            zip_bytes = uploaded.read()
            employees = extract_employee_names(zip_bytes)
            if not employees:
                st.error("No employee folders found inside the zip. "
                         "Make sure the zip has folders like  Timesheet/Leave/EmployeeName/")
            else:
                st.session_state.zip_bytes  = zip_bytes
                st.session_state.employees  = employees
                st.session_state.month      = chosen_month
                st.session_state.year       = int(chosen_year)
                # Preserve any already-entered leave data
                for e in employees:
                    if e not in st.session_state.leaves:
                        st.session_state.leaves[e] = ""
                st.session_state.step = 2
                st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# STEP 2 — Enter leave dates
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 2:
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    mname = month_names[st.session_state.month - 1]

    st.markdown(f"""
    <div class="step-box">
      <h3>🗓️ Step 2 — Enter leave dates for each employee</h3>
      <p>Month: <strong>{mname} {st.session_state.year}</strong> &nbsp;|&nbsp;
         Employees found: <strong>{len(st.session_state.employees)}</strong></p>
    </div>
    """, unsafe_allow_html=True)

    st.info("Type leave dates separated by commas.  Example:  **05/05/2026, 12/05/2026**\n\n"
            "Leave the box empty if the employee took no leaves.")

    updated_leaves = {}
    for emp in st.session_state.employees:
        with st.expander(f"👤  {emp}", expanded=True):
            val = st.text_area(
                label=f"Leave dates for {emp}",
                value=st.session_state.leaves.get(emp, ""),
                placeholder="e.g.  05/05/2026, 12/05/2026",
                key=f"leave_{emp}",
                label_visibility="collapsed",
                height=68
            )
            updated_leaves[emp] = val

            # Live preview of parsed dates
            parsed = parse_leave_dates(val)
            if parsed:
                badges = "  ".join(
                    [f"🔴 {d.strftime('%d %b')}" for d in parsed]
                )
                st.markdown(f"**Leaves added:** {badges}")
            else:
                st.caption("No leaves entered — full month will be 8 hrs/day.")

    col_back, col_next = st.columns(2)
    with col_back:
        if st.button("← Back", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with col_next:
        if st.button("➡️  Review Summary", use_container_width=True, type="primary"):
            st.session_state.leaves = updated_leaves
            st.session_state.step   = 3
            st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# STEP 3 — Review
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 3:
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    mname = month_names[st.session_state.month - 1]
    working_days  = get_working_days(st.session_state.year, st.session_state.month)

    st.markdown(f"""
    <div class="step-box">
      <h3>✅ Step 3 — Review before generating</h3>
      <p>Month: <strong>{mname} {st.session_state.year}</strong> &nbsp;|&nbsp;
         Working days in month: <strong>{len(working_days)}</strong> &nbsp;|&nbsp;
         Max hours: <strong>{len(working_days)*8}</strong></p>
    </div>
    """, unsafe_allow_html=True)

    for emp in st.session_state.employees:
        leave_dates = set(parse_leave_dates(st.session_state.leaves.get(emp, "")))
        total_hours = sum(0 if d in leave_dates else 8 for d in working_days)
        leave_count = len(leave_dates)

        with st.expander(f"👤  {emp}   —   {total_hours} hrs   |   {leave_count} leave(s)", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            col_a.metric("Working Days",  len(working_days))
            col_b.metric("Leaves Taken",  leave_count)
            col_c.metric("Total Hours",   total_hours)

            # Show a small calendar-style table (first 2 weeks as preview)
            import pandas as pd
            day_names = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
            preview_rows = []
            for d in working_days[:10]:
                hrs  = 0 if d in leave_dates else 8
                preview_rows.append({
                    "Date":   d.strftime("%d/%m/%Y"),
                    "Day":    day_names[d.weekday()],
                    "Hours":  hrs,
                    "Status": "Approved"
                })
            if len(working_days) > 10:
                preview_rows.append({
                    "Date": f"… and {len(working_days)-10} more rows",
                    "Day": "", "Hours": "", "Status": ""
                })
            df = pd.DataFrame(preview_rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

    col_back, col_next = st.columns(2)
    with col_back:
        if st.button("← Back to Edit Leaves", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with col_next:
        if st.button("🚀  Generate Excel Files", use_container_width=True, type="primary"):
            st.session_state.step = 4
            st.rerun()

# ═════════════════════════════════════════════════════════════════════════════
# STEP 4 — Download
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 4:
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    mname = month_names[st.session_state.month - 1]
    working_days = get_working_days(st.session_state.year, st.session_state.month)

    st.markdown("""
    <div class="step-box">
      <h3>📥 Step 4 — Download individual Excel files</h3>
      <p>Each file matches the exact column format your portal requires.
         Click each employee's button to download their file.</p>
    </div>
    """, unsafe_allow_html=True)

    st.success(f"✅  Files ready for **{mname} {st.session_state.year}** — "
               f"{len(working_days)} working days per employee.")

    st.write("### Download individual files")
    for emp in st.session_state.employees:
        leave_dates = set(parse_leave_dates(st.session_state.leaves.get(emp, "")))
        total_hours = sum(0 if d in leave_dates else 8 for d in working_days)
        leave_count = len(leave_dates)

        col_info, col_btn = st.columns([3, 1])
        with col_info:
            st.markdown(
                f"<div class='emp-card'>"
                f"<strong>👤 {emp}</strong><br>"
                f"<small style='color:#718096'>"
                f"Total hours: <b>{total_hours}</b> &nbsp;|&nbsp; "
                f"Leaves: <b>{leave_count}</b> &nbsp;|&nbsp; "
                f"Working days: <b>{len(working_days)}</b>"
                f"</small></div>",
                unsafe_allow_html=True
            )
        with col_btn:
            excel_bytes = generate_excel(emp, working_days, leave_dates)
            safe_name   = emp.replace(" ", "_")
            filename    = f"{safe_name}_Timesheet_{mname}_{st.session_state.year}.xlsx"
            st.download_button(
                label    = "⬇️ Download",
                data     = excel_bytes,
                file_name= filename,
                mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key      = f"dl_{emp}",
                use_container_width=True
            )

    st.write("---")
    st.write("### Or download all files in one zip")

    # Build a zip of all Excel files
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for emp in st.session_state.employees:
            leave_dates = set(parse_leave_dates(st.session_state.leaves.get(emp, "")))
            excel_bytes = generate_excel(emp, working_days, leave_dates)
            safe_name   = emp.replace(" ", "_")
            zf.writestr(f"{safe_name}_Timesheet_{mname}_{st.session_state.year}.xlsx", excel_bytes)
    zip_buf.seek(0)

    st.download_button(
        label    = f"⬇️  Download ALL {len(st.session_state.employees)} files as ZIP",
        data     = zip_buf.read(),
        file_name= f"All_Timesheets_{mname}_{st.session_state.year}.zip",
        mime     = "application/zip",
        use_container_width=True,
        type     = "primary"
    )

    st.write("")
    if st.button("🔄  Start Over (new month)", use_container_width=True):
        for k in ("step","zip_bytes","employees","leaves","month","year"):
            st.session_state[k] = None
        st.session_state.step   = 1
        st.session_state.leaves = {}
        st.rerun()
